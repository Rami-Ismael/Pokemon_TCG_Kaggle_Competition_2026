"""Build reports/il_model_deck_selection.xlsx from the merged sweep JSONs.

Sheets:
  Checkpoints    -- provenance for every IL arm (params, data, epochs, offline
                    accuracy/ECE, safetensors sha256) so the model axis is auditable
  Model screen   -- Stage A: each arm vs each pool opponent, win% +/- sigma, aggregate,
                    pooled Glicko + RD + GXE
  Model x Deck   -- Stage B: field win% for each (model, deck) cell
  Familiarity    -- corpus episode count per deck (the (2) audit column)
  Separation     -- which arm pairs are actually separated at 95% CI
  Fallback       -- silent-fallback rate per arm (a non-zero rate means some of
                    the win rate was not produced by the model)
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "scripts"))
# Ladder anchors for the pool: lets the sheet show local Glicko next to the real
# ladder score for the same opponent, which is the whole point of using an
# anchored pool rather than our own agents.
from pool_predictiveness import LADDER_ANCHORS  # noqa: E402

HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="305496")
BOLD = Font(bold=True)
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")

# Corpus episode counts per deck, ace/carry-Pokemon key, train split 2026-07-26
# (4554 episodes). Carried forward from reports/deck_selection.md -- NOT
# recomputed here: the local episode splits are now stubs (24 of 4554 real
# files on disk), so this is the last valid census and re-deriving it would
# require re-streaming the corpus from the HF dataset.
DECK_EPISODES = {
    "marnies_grimmsnarl_ex": 3488,
    "alakazam": 1542,
    "team_rockets_spidops": 784,
    "cynthias_garchomp_ex": 625,
    "mega_kangaskhan_ex": 600,
    "dragapult_ex": 318,
    "mega_lucario_ex": 1,
}

# Provenance per arm, from each checkpoint's train_metadata.json + sha256.
CHECKPOINTS = [
    # arm, ckpt dir, hidden, params, train data, epochs, eval_acc, eval_ece, sha12
    ("il_bc_2ep", "il_agent_2ep_backup", 192, 3_318_721, "train 2026-07-26", 2, 0.74781, None, "758dd7bc55fb"),
    ("il_bc_3ep", "il_agent", 192, 3_318_721, "train 2026-07-26", 3, 0.75344, None, "bce726e56bb2"),
    ("il_bc_4ep", "il_agent_4ep", 192, 3_318_721, "train 2026-07-26", 4, 0.75922, None, "421138c8fe17"),
    ("il_medium_3ep", "il_agent_medium", 320, 10_987_201, "train 2026-07-26", 3, 0.75453, 0.02309, "831244e0cf39"),
    ("il_small_comb_2ep", "il_agent_small_combined", 192, 3_318_721, "combined 07-01+07-26", 2, 0.63656, 0.16304, "bcce8793cf16"),
    ("il_hfstream_comb_3ep", "il_agent_hfstream_combined_3ep", 192, 3_318_721, "combined (HF stream)", 3, 0.75266, 0.02313, "54603f533168"),
    ("il_alldays_3ep", "il_alldays_0804", 192, 3_318_721, "all days (127,748 steps)", 3, 0.75828, 0.01929, "1d67d1acdbb0"),
]

NOTES = {
    "il_bc_2ep": "byte-identical to models/il_agent_winning_827.8 (our best-ever ladder build)",
    "il_bc_3ep": "byte-identical to models/il_agent_3ep; the checkpoint il_agent ships today",
    "il_small_comb_2ep": "ECE 0.163 = 7x worse calibrated than every other arm",
}


def _hdr(ws, row: int, values: list[str]) -> None:
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")


def _autosize(ws, widths: dict[int, int] | None = None) -> None:
    for col in range(1, ws.max_column + 1):
        if widths and col in widths:
            ws.column_dimensions[get_column_letter(col)].width = widths[col]
            continue
        longest = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max(longest + 2, 10), 34)


def sheet_checkpoints(wb: Workbook) -> None:
    ws = wb.create_sheet("Checkpoints")
    ws["A1"] = "IL checkpoint provenance (the model axis)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Deduped by sha256: models/il_agent_3ep is byte-identical to models/il_agent, and "
                "models/il_agent_winning_827.8 to models/il_agent_2ep_backup, so each pair is ONE arm. "
                "models/il_agent_medium_combined is an empty directory and is excluded.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:I2")
    _hdr(ws, 4, ["arm", "checkpoint dir", "hidden", "params", "train data", "epochs",
                 "offline acc", "ECE", "sha256[:12]"])
    r = 5
    for arm, ckpt, hid, params, data, ep, acc, ece, sha in CHECKPOINTS:
        ws.cell(row=r, column=1, value=arm).font = BOLD
        ws.cell(row=r, column=2, value=ckpt)
        ws.cell(row=r, column=3, value=hid)
        ws.cell(row=r, column=4, value=params)
        ws.cell(row=r, column=5, value=data)
        ws.cell(row=r, column=6, value=ep)
        ws.cell(row=r, column=7, value=acc)
        c = ws.cell(row=r, column=8, value=ece if ece is not None else "n/a")
        if ece is not None and ece > 0.05:
            c.fill = WARN_FILL
        ws.cell(row=r, column=9, value=sha)
        if arm in NOTES:
            ws.cell(row=r, column=10, value=NOTES[arm])
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Majority-class baseline on the same eval rows: 0.381").font = BOLD
    _autosize(ws, {5: 26, 10: 60})


def sheet_model_screen(wb: Workbook, merged: dict) -> None:
    ws = wb.create_sheet("Model screen")
    pool = merged["pool"]
    agg = merged["aggregate"]
    ranked = sorted(agg, key=lambda a: -agg[a]["win_pct"])

    ws["A1"] = "Stage A - IL checkpoint screen, deck held FIXED"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Every arm pilots the same deck and faces the same pool. Cells are win% +/- sigma "
                "(sigma = sqrt(p(1-p)/n)). 'Field win%' aggregates all pool games; Glicko is pooled "
                "over the union of all runs from a flat 1500 prior (the standing ladder file was "
                "never touched).")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:L2")

    _hdr(ws, 4, ["arm"] + pool + ["Field win%", "sigma", "95% CI low", "95% CI high",
                                  "Glicko", "RD", "GXE%", "games"])
    r = 5
    for a in ranked:
        ws.cell(row=r, column=1, value=a).font = BOLD
        for i, b in enumerate(pool, start=2):
            cell = merged["per_opponent"][a].get(b)
            ws.cell(row=r, column=i,
                    value=f"{cell['win_pct']:.1f} ± {cell['sigma']:.1f}" if cell else "—")
        g = merged["glicko"][a]
        base = 2 + len(pool)
        ws.cell(row=r, column=base, value=round(agg[a]["win_pct"], 1)).font = BOLD
        ws.cell(row=r, column=base + 1, value=round(agg[a]["sigma"], 1))
        ws.cell(row=r, column=base + 2, value=round(agg[a]["wilson95"][0], 1))
        ws.cell(row=r, column=base + 3, value=round(agg[a]["wilson95"][1], 1))
        ws.cell(row=r, column=base + 4, value=round(g["rating"], 1)).font = BOLD
        ws.cell(row=r, column=base + 5, value=round(g["rd"], 1))
        ws.cell(row=r, column=base + 6, value=round(g["gxe"], 1))
        ws.cell(row=r, column=base + 7, value=agg[a]["games"])
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Pool (reference, same games):").font = BOLD
    r += 1
    _hdr(ws, r, ["opponent", "Glicko", "RD", "GXE%", "ladder anchor"])
    r += 1
    for b in sorted(pool, key=lambda x: -merged["glicko"][x]["rating"]):
        g = merged["glicko"][b]
        ws.cell(row=r, column=1, value=b)
        ws.cell(row=r, column=2, value=round(g["rating"], 1))
        ws.cell(row=r, column=3, value=round(g["rd"], 1))
        ws.cell(row=r, column=4, value=round(g["gxe"], 1))
        anc = LADDER_ANCHORS.get(b)
        ws.cell(row=r, column=5, value=anc[0] if anc else "—")
        r += 1
    _autosize(ws)


def sheet_model_deck(wb: Workbook, cells: dict) -> None:
    """cells: {(model, deck): aggregate-dict}"""
    ws = wb.create_sheet("Model x Deck")
    ws["A1"] = "Stage B - field win% by (checkpoint, deck)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Same pool, same games-per-cell throughout. The 'corpus episodes' row is the (2) "
                "familiarity audit: a deck's win rate is only comparable to another deck's when both "
                "clear the 150-episode floor.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")

    models = sorted({m for m, _ in cells})
    decks = sorted({d for _, d in cells}, key=lambda d: -DECK_EPISODES.get(d, 0))

    _hdr(ws, 4, ["checkpoint \\ deck"] + decks)
    ws.cell(row=5, column=1, value="corpus episodes (07-26)").font = BOLD
    for i, d in enumerate(decks, start=2):
        c = ws.cell(row=5, column=i, value=DECK_EPISODES.get(d, "?"))
        c.font = BOLD
        if DECK_EPISODES.get(d, 0) < 150:
            c.fill = WARN_FILL
    r = 6
    for m in models:
        ws.cell(row=r, column=1, value=m).font = BOLD
        for i, d in enumerate(decks, start=2):
            a = cells.get((m, d))
            if not a:
                ws.cell(row=r, column=i, value="—")
                continue
            ws.cell(row=r, column=i, value=f"{a['win_pct']:.1f} ± {a['sigma']:.1f}")
        r += 1
    r += 1
    ws.cell(row=r, column=1,
            value="Red = below the 150-episode familiarity floor: unmeasured, not bad.").font = BOLD
    _autosize(ws, {1: 26})


def sheet_familiarity(wb: Workbook) -> None:
    ws = wb.create_sheet("Familiarity")
    ws["A1"] = "(2) Familiarity audit - corpus episodes per deck"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Train split 2026-07-26 (4554 episodes), deck identity = ace/carry Pokemon. "
                "Carried forward from reports/deck_selection.md; the local splits are now stubs "
                "(24 of 4554 real files), so this could not be re-derived without re-streaming "
                "the HF corpus. NOTE: arms trained on additional days (il_alldays_3ep, "
                "il_hfstream_comb_3ep, il_small_comb_2ep) saw MORE episodes of these decks than "
                "this column shows - their true familiarity is UNMEASURED.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    _hdr(ws, 4, ["deck (ace/carry)", "episodes", "clears 150 floor?", "in benchmark?"])
    r = 5
    for d, n in sorted(DECK_EPISODES.items(), key=lambda kv: -kv[1]):
        ws.cell(row=r, column=1, value=d)
        ws.cell(row=r, column=2, value=n)
        c = ws.cell(row=r, column=3, value="yes" if n >= 150 else "NO - unmeasured")
        c.fill = GOOD_FILL if n >= 150 else WARN_FILL
        r += 1
    _autosize(ws, {1: 28, 2: 12, 3: 20, 4: 16})


def sheet_separation(wb: Workbook, merged: dict, name: str, title: str) -> None:
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Two arms are only called different when their aggregate Wilson 95% CIs do not "
                "overlap. Everything marked NOT SEPARATED is a tie at this sample size, regardless "
                "of the point-estimate gap.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E2")
    _hdr(ws, 4, ["stronger arm", "weaker arm", "gap (pp)", "separated at 95%?"])
    r = 5
    for s in merged["separation"]:
        ws.cell(row=r, column=1, value=s["a"])
        ws.cell(row=r, column=2, value=s["b"])
        ws.cell(row=r, column=3, value=round(s["gap_pp"], 1))
        c = ws.cell(row=r, column=4, value="SEPARATED" if s["separated"] else "not separated")
        c.fill = GOOD_FILL if s["separated"] else WARN_FILL
        r += 1
    _autosize(ws)


def sheet_fallback(wb: Workbook, fallbacks: dict) -> None:
    ws = wb.create_sheet("Fallback")
    ws["A1"] = "Silent-fallback rate per arm"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("il_agent catches its own errors and plays a non-ML _safe_choice instead. A non-zero "
                "rate means that share of decisions was NOT made by the checkpoint under test, so "
                "the win rate is partly a measurement of the fallback heuristic.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E2")
    _hdr(ws, 4, ["arm", "decisions", "fallbacks", "fallback rate %", "first reason"])
    r = 5
    for arm, d in sorted(fallbacks.items()):
        ws.cell(row=r, column=1, value=arm)
        ws.cell(row=r, column=2, value=d.get("decisions", 0))
        ws.cell(row=r, column=3, value=d.get("fallbacks", 0))
        rate = d.get("rate_pct", 0.0)
        c = ws.cell(row=r, column=4, value=round(rate, 3))
        c.fill = WARN_FILL if rate > 1.0 else GOOD_FILL
        ws.cell(row=r, column=5, value=d.get("first_reason") or "—")
        r += 1
    _autosize(ws, {5: 40})


def sheet_leaderboard(wb: Workbook, merged: dict) -> None:
    ws = wb.create_sheet("Local leaderboard")
    ws["A1"] = "Local leaderboard - every agent on one pooled Glicko scale"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Challengers never play each other; they are linked through the shared 8-agent pool. "
                "Two agents are only different when rating +/- 1.96*RD does not overlap. The 'ladder' "
                "column is the agent's REAL Kaggle score where one is known -- compare it to the local "
                "rating, not to other local ratings.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:G2")
    _hdr(ws, 4, ["#", "agent", "Glicko", "RD", "95% CI", "GXE%", "role", "real ladder"])
    g = merged["glicko"]
    pool = set(merged["pool"])
    r = 5
    for i, n in enumerate(sorted(g, key=lambda x: -g[x]["rating"]), 1):
        d = g[n]
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=n).font = BOLD
        ws.cell(row=r, column=3, value=round(d["rating"], 1)).font = BOLD
        ws.cell(row=r, column=4, value=round(d["rd"], 0))
        lo, hi = d["rating"] - 1.96 * d["rd"], d["rating"] + 1.96 * d["rd"]
        ws.cell(row=r, column=5, value=f"[{lo:.0f}, {hi:.0f}]")
        ws.cell(row=r, column=6, value=round(d["gxe"], 1))
        ws.cell(row=r, column=7, value="pool" if n in pool else "challenger")
        anc = LADDER_ANCHORS.get(n)
        c = ws.cell(row=r, column=8, value=anc[0] if anc else "—")
        if anc:
            c.fill = GOOD_FILL
        r += 1
    r += 1
    ws.cell(row=r, column=1,
            value="Spearman rho vs ladder: +0.765 (n=18, p=0.0004) on ALL anchors -- but 14 of 18 are self-reported by their authors, NOT verified. On the 4 we read ourselves: rho = +1.000 (n=4, p=0.087). agent_core_improved corrected 804.0 -> 685.3 (mean of 4 settled same-build resubmits)."
            ).font = BOLD
    _autosize(ws, {2: 44, 5: 18})


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--stage-a", type=Path, required=True)
    ap.add_argument("--stage-b", type=Path, default=None)
    ap.add_argument("--fallbacks", type=Path, default=None)
    ap.add_argument("--leaderboard", type=Path, default=None,
                    help="merged JSON covering ALL agents (IL arms + our other agents + pool)")
    ap.add_argument("--out", type=Path, required=True)
    args = ap.parse_args()

    merged_a = json.loads(args.stage_a.read_text())

    wb = Workbook()
    wb.remove(wb.active)
    if args.leaderboard and args.leaderboard.exists():
        sheet_leaderboard(wb, json.loads(args.leaderboard.read_text()))
    sheet_checkpoints(wb)
    sheet_model_screen(wb, merged_a)

    cells: dict[tuple[str, str], dict] = {}
    for label, a in merged_a["aggregate"].items():
        m, _, d = label.partition("@")
        cells[(m, d)] = a
    if args.stage_b and args.stage_b.exists():
        merged_b = json.loads(args.stage_b.read_text())
        for label, a in merged_b["aggregate"].items():
            m, _, d = label.partition("@")
            cells[(m, d)] = a
        sheet_separation(wb, merged_b, "Separation (deck)",
                         "Stage B - which (checkpoint, deck) cells are separated?")
    sheet_model_deck(wb, cells)
    sheet_familiarity(wb)
    sheet_separation(wb, merged_a, "Separation (model)",
                     "Stage A - which checkpoint pairs are separated, deck held fixed?")
    if args.fallbacks and args.fallbacks.exists():
        sheet_fallback(wb, json.loads(args.fallbacks.read_text()))

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"saved: {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
